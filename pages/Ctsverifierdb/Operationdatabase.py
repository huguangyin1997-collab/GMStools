# Operationdatabase.py
import sqlite3
import os
from datetime import datetime

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class DatabaseExporter:
    """数据库导出工具，将 results 表导出为 Excel"""

    @staticmethod
    def hex_preview(data):
        """将 BLOB 转换为十六进制字符串；若为字符串则原样返回"""
        if data is None:
            return "NULL"
        if isinstance(data, bytes):
            if len(data) == 0:
                return "<empty>"
            hex_str = data.hex()
            spaced_hex = ' '.join(hex_str[i:i+2] for i in range(0, len(hex_str), 2))
            return spaced_hex
        elif isinstance(data, str):
            return data if data else "<empty>"
        else:
            return str(data)

    @staticmethod
    def export_results_to_excel(db_file, output_filename=None):
        """
        将指定的 SQLite 数据库中的 results 表导出为 Excel 文件
        :param db_file: 数据库文件路径
        :param output_filename: 输出文件名，默认为 None 则使用 "db_new.xlsx"
        :return: (success, message, output_file)
        """
        if not OPENPYXL_AVAILABLE:
            return False, "缺少 openpyxl 库，请执行: pip install openpyxl", None

        if not os.path.isfile(db_file):
            return False, f"数据库文件不存在: {db_file}", None

        try:
            conn = sqlite3.connect(db_file)
            cursor = conn.cursor()

            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='results'")
            if not cursor.fetchone():
                conn.close()
                return False, "数据库中没有 'results' 表。", None

            cursor.execute("SELECT * FROM results")
            rows = cursor.fetchall()
            if not rows:
                conn.close()
                return False, "表为空。", None

            col_names = [desc[0] for desc in cursor.description]

            processed_rows = []
            for row in rows:
                processed_row = []
                for i, val in enumerate(row):
                    if col_names[i] in ('testmetrics', 'testresulthistory', 'testscreenshotsmetadata'):
                        processed_row.append(DatabaseExporter.hex_preview(val))
                    else:
                        processed_row.append(str(val) if val is not None else "NULL")
                processed_rows.append(processed_row)

            wb = Workbook()
            ws = wb.active
            ws.title = "results"

            ws.append(col_names)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            for row_data in processed_rows:
                ws.append(row_data)

            for col_idx, _ in enumerate(col_names, 1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    cell_value = row[0].value
                    if cell_value:
                        max_len = max(max_len, len(str(cell_value)))
                adjusted_width = min(max_len + 2, 80)
                ws.column_dimensions[col_letter].width = adjusted_width

            base_dir = os.path.dirname(db_file)
            if output_filename is None:
                output_filename = "db_new.xlsx"
            out_path = os.path.join(base_dir, output_filename)

            wb.save(out_path)
            conn.close()
            return True, f"成功导出到 {out_path}", out_path

        except sqlite3.Error as e:
            return False, f"数据库错误: {e}", None
        except Exception as e:
            return False, f"导出失败: {e}", None


class DatabaseImporter:
    """从 Excel 恢复数据库（对比更新）"""

    @staticmethod
    def import_excel_to_db(excel_file, output_db):
        """直接替换方式（保留原功能）"""
        if not OPENPYXL_AVAILABLE:
            return False, "缺少 openpyxl 库，请执行: pip install openpyxl"

        try:
            wb = load_workbook(excel_file)
            ws = wb.active

            headers = [cell.value for cell in ws[1]]
            data_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                data_rows.append(row)

            conn = sqlite3.connect(output_db)
            cursor = conn.cursor()

            col_defs = [f'"{col}" TEXT' for col in headers]
            create_sql = f'CREATE TABLE results ({", ".join(col_defs)})'
            cursor.execute(create_sql)

            placeholders = ','.join(['?'] * len(headers))
            insert_sql = f'INSERT INTO results VALUES ({placeholders})'
            for row in data_rows:
                cursor.execute(insert_sql, row)

            conn.commit()
            conn.close()
            return True, f"成功恢复数据库: {output_db}"
        except Exception as e:
            return False, f"恢复失败: {e}"

    @staticmethod
    def compare_and_update(new_excel, old_excel, target_db):
        """
        对比两个 Excel 文件，将差异更新到目标数据库
        :param new_excel: 新 Excel 文件路径
        :param old_excel: 旧 Excel 文件路径
        :param target_db: 目标数据库文件路径
        :return: (success, message)
        """
        if not OPENPYXL_AVAILABLE:
            return False, "缺少 openpyxl 库，请执行: pip install openpyxl"

        if not os.path.isfile(new_excel):
            return False, f"新 Excel 文件不存在: {new_excel}"
        if not os.path.isfile(old_excel):
            return False, f"旧 Excel 文件不存在: {old_excel}"
        if not os.path.isfile(target_db):
            return False, f"目标数据库文件不存在: {target_db}"

        try:
            wb_new = load_workbook(new_excel)
            ws_new = wb_new.active
            headers = [cell.value for cell in ws_new[1]]
            pk_col = '_id' if '_id' in headers else None

            new_data = {}
            for row in ws_new.iter_rows(min_row=2, values_only=True):
                if pk_col:
                    pk_idx = headers.index(pk_col)
                    pk = row[pk_idx]
                    new_data[pk] = dict(zip(headers, row))
                else:
                    new_data[tuple(row)] = dict(zip(headers, row))

            wb_old = load_workbook(old_excel)
            ws_old = wb_old.active
            old_data = {}
            for row in ws_old.iter_rows(min_row=2, values_only=True):
                if pk_col:
                    pk_idx = headers.index(pk_col)
                    pk = row[pk_idx]
                    old_data[pk] = dict(zip(headers, row))
                else:
                    old_data[tuple(row)] = dict(zip(headers, row))

            if pk_col:
                new_keys = set(new_data.keys())
                old_keys = set(old_data.keys())

                inserted = new_keys - old_keys
                deleted = old_keys - new_keys
                updated = {k for k in new_keys & old_keys if new_data[k] != old_data[k]}

                conn = sqlite3.connect(target_db)
                cursor = conn.cursor()

                for pk in deleted:
                    cursor.execute(f"DELETE FROM results WHERE {pk_col}=?", (pk,))

                col_names = headers
                placeholders = ','.join(['?'] * len(col_names))
                insert_sql = f'INSERT OR REPLACE INTO results ({",".join(col_names)}) VALUES ({placeholders})'

                for pk in inserted | updated:
                    row_dict = new_data[pk]
                    row_values = [row_dict[col] for col in col_names]
                    cursor.execute(insert_sql, row_values)

                conn.commit()
                conn.close()

                msg = f"更新完成: 新增 {len(inserted)} 条, 修改 {len(updated)} 条, 删除 {len(deleted)} 条"
            else:
                conn = sqlite3.connect(target_db)
                cursor = conn.cursor()
                cursor.execute("DELETE FROM results")
                col_names = headers
                placeholders = ','.join(['?'] * len(col_names))
                insert_sql = f'INSERT INTO results ({",".join(col_names)}) VALUES ({placeholders})'
                for row_dict in new_data.values():
                    row_values = [row_dict[col] for col in col_names]
                    cursor.execute(insert_sql, row_values)
                conn.commit()
                conn.close()
                msg = f"无主键，已完全替换表，共 {len(new_data)} 条记录"

            return True, msg

        except Exception as e:
            return False, f"对比更新失败: {e}"